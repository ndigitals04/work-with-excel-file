import openpyxl
workbook = openpyxl.load_workbook("C:\\Users\\USER\\Downloads\\product-list.xlsx")
product_sheet = workbook["product-list"]
product_list = []
for row in range(2, product_sheet.max_row + 1):
    product_list.append(product_sheet.cell(row,2).value)
print("Appending to list completed")
product_file = open("C:/users/user/desktop/product_list.txt", "w")
# product_sup_file = open("C:/users/user/desktop/product_Supermarket_list.txt", "w")
drug_workbook = openpyxl.load_workbook("C:\\Users\\USER\\Desktop\\product-drug-list.xlsx")
drug_sheet = drug_workbook["drug-list"]

sup_workbook = openpyxl.load_workbook("C:\\Users\\USER\\Desktop\\product-supermarket-list.xlsx")
sup_sheet = sup_workbook["supermarket-list"]

drug_sheet.append(["S/N", "Product Name", "Department"])
sup_sheet.append(["S/N", "Product Name", "Department"])

for i  in range(len(product_list)):
    if product_list[i] == None:
        pass
    else:
        while True:
            print(product_list[i])
            department = input("Select department with 'p' or 's' or 'r': ")
            if department == "s":
                sup_sheet.append([i+1,product_list[i], "supermarket"])
                sup_workbook.save("C:\\Users\\USER\\Desktop\\product-supermarket-list.xlsx")
                break
            elif department == "p":
                drug_sheet.append([i+1,product_list[i], "pharmacy"])
                drug_workbook.save("C:\\Users\\USER\\Desktop\\product-drug-list.xlsx")
                break
            elif department == "r":
                break
            else:
                print("Accepted letters are 's' or 'p' ")

print("Done")
