import openpyxl
workbook = openpyxl.load_workbook("C:\\Users\\USER\\Desktop\\product-supermarket-list.xlsx")
product_sheet = workbook["supermarket-list"]
product_list = []
for row in range(2, product_sheet.max_row + 1):
    product_list.append(product_sheet.cell(row,2).value)
print("Appending to list completed")
product_file = open("C:/users/user/desktop/product-supermarket-list.txt", "w")
for product in product_list:
    if product == None:
        pass
    else:
        product_file.write(product + "\n")

product_file.close()